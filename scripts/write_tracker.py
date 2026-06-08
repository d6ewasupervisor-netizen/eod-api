import argparse
import json
import os
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


K_COLUMN = 11
L_COLUMN = 12


def owner_lock_path(workbook_path):
    return workbook_path.with_name(f"~${workbook_path.name}")


def assert_not_locked(workbook_path):
    lock_path = owner_lock_path(workbook_path)
    if lock_path.exists():
        raise PermissionError(f"Workbook appears to be open or locked: {lock_path}")
    try:
        with open(workbook_path, "r+b"):
            pass
    except PermissionError as exc:
        raise PermissionError(f"Workbook is locked or not writable: {workbook_path}") from exc


def backup_path_for(workbook_path):
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return workbook_path.with_name(f"{workbook_path.stem}.bak-{stamp}{workbook_path.suffix}")


def parse_rows():
    payload = json.loads(sys.stdin.read() or "{}")
    rows = payload.get("rows", payload if isinstance(payload, list) else [])
    if not isinstance(rows, list):
        raise ValueError("Expected JSON array or object with rows array.")
    out = []
    for row in rows:
        row_index = int(row["rowIndex"])
        if row_index < 1:
            raise ValueError(f"Invalid rowIndex: {row_index}")
        out.append({
            "rowIndex": row_index,
            "K": "" if row.get("K") is None else str(row.get("K")),
            "L": "" if row.get("L") is None else str(row.get("L")),
        })
    return out


def main():
    parser = argparse.ArgumentParser(description="Write approved K/L tracker cells with a backup.")
    parser.add_argument("workbook_path")
    parser.add_argument("sheet_name")
    args = parser.parse_args()

    workbook_path = Path(args.workbook_path)
    try:
        if not workbook_path.exists():
            raise FileNotFoundError(f"Workbook not found: {workbook_path}")
        rows = parse_rows()
        assert_not_locked(workbook_path)
        backup_path = backup_path_for(workbook_path)
        shutil.copy2(workbook_path, backup_path)

        keep_vba = workbook_path.suffix.lower() == ".xlsm"
        workbook = load_workbook(workbook_path, keep_vba=keep_vba, data_only=False)
        if args.sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {args.sheet_name}")
        sheet = workbook[args.sheet_name]

        written = []
        for row in rows:
            row_index = row["rowIndex"]
            sheet.cell(row=row_index, column=K_COLUMN).value = row["K"]
            sheet.cell(row=row_index, column=L_COLUMN).value = row["L"]
            written.append(row_index)

        workbook.save(workbook_path)
        print(json.dumps({
            "backupPath": str(backup_path),
            "written": written,
            "skipped": [],
            "errors": [],
        }, ensure_ascii=False))
        return 0
    except Exception as exc:
        print(str(exc), file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
