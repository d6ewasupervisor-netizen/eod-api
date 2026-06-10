import argparse
import json
import sys

from openpyxl import load_workbook


def cell_value(row, index):
    if index - 1 >= len(row):
        return ""
    value = row[index - 1]
    return "" if value is None else value


def string_value(value):
    if value is None:
        return ""
    return str(value).strip()


def is_data_row(payload):
    return any(string_value(payload.get(key)) for key in ("store", "categoryNumber", "pogId", "setType", "currentK", "currentL"))


def main():
    parser = argparse.ArgumentParser(description="Read tracker workbook rows as JSON.")
    parser.add_argument("workbook_path")
    parser.add_argument("sheet_name")
    parser.add_argument("--header-row", type=int, default=1)
    parser.add_argument("--max-empty-rows", type=int, default=25)
    args = parser.parse_args()

    try:
        workbook = load_workbook(args.workbook_path, read_only=True, data_only=False)
        if args.sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {args.sheet_name}")
        sheet = workbook[args.sheet_name]
        rows = []
        empty_run = 0
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_index <= args.header_row:
                continue
            payload = {
                "rowIndex": row_index,
                "store": string_value(cell_value(row, 3)),
                "categoryNumber": string_value(cell_value(row, 4)),
                "pogName": string_value(cell_value(row, 5)),
                "pogId": string_value(cell_value(row, 8)),
                "setType": string_value(cell_value(row, 9)),
                "currentK": string_value(cell_value(row, 11)),
                "currentL": string_value(cell_value(row, 12)),
            }
            if not is_data_row(payload):
                empty_run += 1
                if empty_run >= args.max_empty_rows:
                    break
                continue
            empty_run = 0
            rows.append(payload)
        print(json.dumps(rows, ensure_ascii=False))
        return 0
    except Exception as exc:
        print(str(exc), file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
