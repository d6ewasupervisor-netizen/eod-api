"""
D1 Outstanding Sets – Workbook Annotator

Reads the workbook update JSON produced by d1-outstanding-si-reconcile.js
and writes Supervisor Comment annotations into a copy of the workbook.

Usage:
    python scripts/d1-outstanding-update-workbook.py <path-to-workbook-updates.json>
    python scripts/d1-outstanding-update-workbook.py <path-to-workbook-updates.json> --workbook "path/to/workbook.xlsx"
"""

import sys
import json
import os
import shutil
import datetime
import openpyxl

WORKBOOK_PATH = r'C:\Users\tgaut\Downloads\Outstanding Sets In District 1.xlsx'
AUTHOR_TAG = 'TAG'
TODAY_STR = datetime.date.today().strftime('%m/%d/%y')  # e.g. "06/26/26"


def versioned_path(desired_path):
    """Return a non-colliding path using file-utils versioning scheme."""
    if not os.path.exists(desired_path):
        return desired_path
    base, ext = os.path.splitext(desired_path)
    v = 2
    while True:
        candidate = f'{base} version {v}{ext}'
        if not os.path.exists(candidate):
            return candidate
        v += 1


def parse_args():
    args = sys.argv[1:]
    updates_path = None
    workbook_path = WORKBOOK_PATH
    replace_pattern = None   # if set, replace matching segment in existing comment instead of appending

    i = 0
    while i < len(args):
        if args[i] == '--workbook' and i + 1 < len(args):
            workbook_path = args[i + 1]
            i += 2
        elif args[i] == '--replace-pattern' and i + 1 < len(args):
            replace_pattern = args[i + 1]
            i += 2
        else:
            updates_path = args[i]
            i += 1

    if not updates_path:
        print('Usage: python scripts/d1-outstanding-update-workbook.py <updates.json> [--workbook path.xlsx] [--replace-pattern "regex"]')
        sys.exit(1)

    return updates_path, workbook_path, replace_pattern


def load_updates(updates_path):
    with open(updates_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def main():
    updates_path, workbook_src, replace_pattern = parse_args()

    print(f'[wb-updater] Updates:  {updates_path}')
    print(f'[wb-updater] Workbook: {workbook_src}')

    updates = load_updates(updates_path)
    if not updates:
        print('[wb-updater] No updates to apply.')
        return

    # Build lookup: dbkey → note (some DBKeys appear once, some might be duplicated if same POG different week)
    # Use (dbkey, week, store) as unique key
    update_map = {}
    for u in updates:
        key = (str(u['dbkey']), str(u.get('week', '')), str(u.get('store', '')))
        update_map[key] = u

    print(f'[wb-updater] {len(update_map)} annotations to apply')

    # Determine output path
    src_base = os.path.splitext(workbook_src)[0]
    src_ext = os.path.splitext(workbook_src)[1]
    out_dir = os.path.dirname(workbook_src)
    today_tag = datetime.date.today().strftime('%Y-%m-%d')
    desired_copy = os.path.join(out_dir, f'Outstanding Sets In District 1 - Annotated {today_tag}{src_ext}')
    out_path = versioned_path(desired_copy)

    # Copy workbook to output path
    print(f'[wb-updater] Copying to: {out_path}')
    shutil.copy2(workbook_src, out_path)

    # Open the copy
    wb = openpyxl.load_workbook(out_path)

    if 'Sets' not in wb.sheetnames:
        print(f'[wb-updater] ERROR: no "Sets" sheet in workbook')
        sys.exit(1)

    ws = wb['Sets']

    # Find column indices from header row
    headers = [str(c.value).strip() if c.value is not None else '' for c in ws[1]]
    print(f'[wb-updater] Headers: {headers}')

    try:
        col_week = headers.index('Week') + 1          # 1-based
        col_store = headers.index('Store') + 1
        col_dbkey = headers.index('POG DBKey') + 1
        col_comment = headers.index('Supervisor Comments') + 1
    except ValueError as e:
        print(f'[wb-updater] ERROR: could not find header column: {e}')
        print(f'[wb-updater] Available headers: {headers}')
        sys.exit(1)

    print(f'[wb-updater] Columns: week={col_week} store={col_store} dbkey={col_dbkey} comment={col_comment}')

    applied = 0
    skipped_201 = 0
    not_matched = 0

    # Row 1 is header, data starts at row 2
    for row in ws.iter_rows(min_row=2):
        week_cell = row[col_week - 1]
        store_cell = row[col_store - 1]
        dbkey_cell = row[col_dbkey - 1]
        comment_cell = row[col_comment - 1]

        # Check category column (index 4 = column E = index 4 in 0-based row tuple)
        cat_col_idx = 4  # 0-based, column 5 = Commodity
        cat_val = str(row[cat_col_idx].value) if row[cat_col_idx].value is not None else ''
        if '201 CANDY' in cat_val:
            skipped_201 += 1
            continue

        dbkey_val = str(dbkey_cell.value) if dbkey_cell.value is not None else ''
        week_val = str(week_cell.value) if week_cell.value is not None else ''
        store_val = str(store_cell.value) if store_cell.value is not None else ''

        lookup = (dbkey_val, week_val, store_val)
        if lookup not in update_map:
            not_matched += 1
            continue

        update = update_map[lookup]
        new_note = update.get('workbookNote', '')
        if not new_note:
            not_matched += 1
            continue

        # Update comment: replace matching segment or append
        existing = str(comment_cell.value).strip() if comment_cell.value is not None else ''
        if replace_pattern and existing:
            import re
            # Strip any segments matching the pattern (semi-colon separated)
            parts = [p.strip() for p in existing.split(';')]
            parts = [p for p in parts if p and not re.search(replace_pattern, p, re.IGNORECASE)]
            parts.append(new_note)
            comment_cell.value = '; '.join(parts)
        elif existing:
            comment_cell.value = f'{existing}; {new_note}'
        else:
            comment_cell.value = new_note

        print(f'  [wb-updater] Row store={store_val} week={week_val} dbkey={dbkey_val}: "{comment_cell.value}"')
        applied += 1

    wb.save(out_path)
    wb.close()

    print('')
    print(f'[wb-updater] Done.')
    print(f'  Applied:        {applied}')
    print(f'  Skipped (201):  {skipped_201}')
    print(f'  Not matched:    {not_matched}')
    print(f'  Output:         {out_path}')


if __name__ == '__main__':
    main()
